import pandas as pd
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

#fecha actual
fecha_actual = datetime.now()
fecha_venci = datetime.now().date()
fecha = fecha_actual.strftime("%d-%m-%Y")

#Ruta archivo final excel
archivo_final_excel= f"C:/Users/Ricardo Sarda/Desktop/MM/Pólizas/Pólizas {fecha}.xlsx"

#archivos
metabase = "C:/Users/Ricardo Sarda/Downloads/raw_data_finance_purchases_spain_2024-12-17T08_38_38.043797Z.xlsx" #metabase
Santander = "C:/Users/Ricardo Sarda/Downloads/SF_EstadosPropuesta_2024_12_17_9_38.xlsx" #santander
Sabadell = "C:/Users/Ricardo Sarda/Downloads/MainServlet - 2024-12-17T093805.612.xls" #sabadell
Sofinco = "C:/Users/Ricardo Sarda/Downloads/ExportListDraws - 2024-12-16T093944.618.xlsx" #sofinco
metabase = pd.read_excel(metabase)
metabase['purchase_date'] = pd.to_datetime(metabase['purchase_date'], format='%d/%m/%Y',errors='coerce', dayfirst=True)
metabase['registration_date'] = pd.to_datetime(metabase['registration_date'], format='%d/%m/%Y')
Santander = pd.read_excel(Santander)
Santander['Fecha Vencimiento'] = pd.to_datetime(Santander['Fecha Vencimiento'], format='%d/%m/%Y')
Santander = Santander.sort_values(by='Fecha Vencimiento',ascending=True)
Santander['license_plate'] = Santander['Matrícula']
Santander['Bastiror '] = Santander['Bastidor '].str.upper()
Santander['Importe Documentación'] = Santander['Importe Documentación'].astype(float)
Santanderp = Santander.copy()
Sabadell = pd.read_excel(Sabadell, engine='xlrd')
Sabadell['Fecha Vencimiento'] = pd.to_datetime(Sabadell['Fecha Vencimiento'], format='%d/%m/%Y')
Sabadell['license_plate'] = Sabadell['Matrícula']
Sabadell = Sabadell.loc[Sabadell['Estado'].isin(['Financed'])]
Sabadellp = Sabadell.copy()
Sofinco = pd.read_excel(Sofinco)
Sofinco['End date'] = pd.to_datetime(Sofinco['End date'], format='%d/%m/%Y')
Sofinco['Bastidor'] = Sofinco['VIN']
Sofinco['Estado'] = Sofinco['Phase']
Sofincop = Sofinco.copy()

#rescates Santander
#seleccionar vencimientos
mask = (Santander['Fecha Vencimiento'].dt.date == fecha_venci) | \
       (Santander['Fecha Vencimiento'].dt.date == (fecha_venci + timedelta(days=1))) | \
       (Santander['Fecha Vencimiento'].dt.date == (fecha_venci + timedelta(days=2))) | \
       (Santander['Fecha Vencimiento'].dt.date == (fecha_venci + timedelta(days=3))) | \
       (Santander['Fecha Vencimiento'].dt.date == (fecha_venci + timedelta(days=4))) | \
       (Santander['Fecha Vencimiento'].dt.date == (fecha_venci + timedelta(days=5)))

vencimientos = Santander[mask]
vencimiemtos_columnas = ['license_plate', 'Estado','Bastidor ','Fecha Vencimiento', 'Importe Documentación']
vencimientos = vencimientos[vencimiemtos_columnas]

#limpieza de pólizas normales
Santanderp = Santanderp.merge(metabase[['license_plate','actual_credit_policy']],left_on='license_plate', right_on='license_plate', how='left')
Sabadellp = Sabadellp.merge(metabase[['license_plate','actual_credit_policy']],left_on='license_plate', right_on='license_plate', how='left')
Sofincop = Sofincop.merge(metabase[['frame_number','actual_credit_policy']],left_on='Bastidor', right_on='frame_number', how='left')
columnas_santander = ['Nº póliza', 'Matrícula','Bastidor ','Marca/Modelo', 'Importe Documentación','Fecha Entrada Stock desde', 'Fecha Vencimiento', 'Estado', 'Estado de la Ficha Téc.','actual_credit_policy']
columnas_sabadell = ['Contrato', 'Bastidor', 'Matrícula','Marca', 'Modelo', 'Linea', 'Fecha Inicio', 'Fecha Vencimiento','Estado', 'Importe Financiado', 'Capital Pendiente', 'Contrato Recibido', 'actual_credit_policy']
columnas_sofinco = ['Contract', 'Phase', 'Financial plan', 'Asset type', 'Make', 'Invoice', 'VIN',  'Start date', 'End date','Amount', 'Estado', 'actual_credit_policy']
Santanderp = Santanderp[columnas_santander]
Sabadellp = Sabadellp[columnas_sabadell]
Sabadellp = Sabadellp.drop_duplicates(subset=['Bastidor'], keep='first')
Sofincop = Sofincop[columnas_sofinco]

#rescate santander
rescate_santander = metabase.loc[metabase['actual_credit_policy'].isin(['santanderSales'])]
rescate_santander = rescate_santander.loc[rescate_santander['stock_status'].isin(['sold'])]
rescate_santander = rescate_santander.loc[rescate_santander['productive_status'].isin(['delivered','readyToDeliver',])]
rescate_santander = rescate_santander.merge(Santander[['license_plate', 'Bastidor ']], left_on='license_plate', right_on='license_plate', how='left')
rescate_santander = rescate_santander.merge(Santander[['license_plate', 'Importe Documentación']], left_on='license_plate', right_on='license_plate', how='left')
rescate_santander = rescate_santander.merge(Santander[['Matrícula', 'Fecha Vencimiento']], left_on='license_plate', right_on='Matrícula', how='left')
rescate_santander = pd.concat([rescate_santander, vencimientos], axis=0)
rescate_santander = rescate_santander.merge(Santander[['license_plate', 'Estado']],left_on='license_plate', right_on='license_plate', how='left')
rescate_santander['Estado'] = rescate_santander['Estado_y']
rescate_santander['Estado'] = rescate_santander['Estado'].fillna('Fuera de póliza')
rescate_santander_columnas = ['license_plate', 'stock_status', 'productive_status','Bastidor ','Importe Documentación','Estado','Fecha Vencimiento']
rescate_santander = rescate_santander[rescate_santander_columnas]
rescate_santander['stockapp'] = rescate_santander.apply(lambda row: f"{row['license_plate']}#",axis=1)
rescate_santander = rescate_santander.sort_values(by='Fecha Vencimiento',ascending=True)
rescate_santander = rescate_santander.drop_duplicates(subset=['license_plate'], keep='first')

#rescate sabadell
rescate_sabadell = metabase.loc[metabase['actual_credit_policy'].isin(['sabadellSales'])]
rescate_sabadell = rescate_sabadell.loc[rescate_sabadell['stock_status'].isin(['sold'])]
rescate_sabadell = rescate_sabadell.loc[rescate_sabadell['productive_status'].isin(['delivered','readyToDeliver',])]
rescate_sabadell = rescate_sabadell.merge(Sabadell[['license_plate', 'Estado']], left_on='license_plate', right_on='license_plate', how='left')
rescate_sabadell = rescate_sabadell.merge(Sabadell[['license_plate', 'Bastidor']], left_on='license_plate', right_on='license_plate', how='left')
rescate_sabadell = rescate_sabadell.merge(Sabadell[['license_plate', 'Importe Financiado']], left_on='license_plate', right_on='license_plate', how='left')
rescate_sabadell['stockapp'] = rescate_sabadell.apply(lambda row: f"{row['license_plate']}#",axis=1)
rescate_sabadell_columnas = ['license_plate', 'stock_status', 'productive_status','Bastidor','Importe Financiado','Estado','stockapp']
rescate_sabadell = rescate_sabadell[rescate_sabadell_columnas]
rescate_sabadell['Estado'] = rescate_sabadell['Estado'].fillna('Fuera de póliza')
rescate_sabadell = rescate_sabadell.sort_values(by='Estado',ascending=True)

#rescate sofinco
rescate_sofinco = metabase.loc[metabase['actual_credit_policy'].isin(['sofincoSales'])]
rescate_sofinco = rescate_sofinco.loc[rescate_sofinco['stock_status'].isin(['sold'])]
rescate_sofinco = rescate_sofinco.loc[rescate_sofinco['productive_status'].isin(['delivered','readyToDeliver',])]
rescate_sofinco = rescate_sofinco.merge(Sofinco[['Bastidor', 'Estado']], left_on='frame_number', right_on='Bastidor', how='left')
rescate_sofinco = rescate_sofinco.merge(Sofinco[['Bastidor', 'Amount']], left_on='frame_number', right_on='Bastidor', how='left')
rescate_sofinco['stockapp'] = rescate_sofinco.apply(lambda row: f"{row['license_plate']}#",axis=1)
rescate_sofinco_columnas = ['license_plate', 'stock_status', 'productive_status','frame_number','Amount','Estado','stockapp']
rescate_sofinco = rescate_sofinco[rescate_sofinco_columnas]
rescate_sofinco['Estado'] = rescate_sofinco['Estado'].fillna('Fuera de póliza')
rescate_sofinco = rescate_sofinco.sort_values(by='Estado',ascending=True)

#motossinpoliza
motosparadoc = metabase.loc[metabase['status'].isin(['pricing','quality','refurbish'])]
motosparadoc = motosparadoc.loc[motosparadoc['actual_credit_policy'].isnull()]
extrasdesab = motosparadoc.loc[motosparadoc['santandersales'].isin(['santanderSales'])]

#motos libres para Santander
motosparsantander = motosparadoc.loc[motosparadoc['santandersales'].isnull()]
motosparsantander = motosparsantander.loc[motosparsantander['wavi'].isnull()]
motosparsantander = motosparsantander.loc[motosparsantander['santanderrenting'].isnull()]
motosparsantander = motosparsantander.loc[motosparsantander['sabadellsales'].isnull()]
motosparsantander = motosparsantander.loc[motosparsantander['sofincosales'].isnull()]
motosparsantander = motosparsantander.loc[motosparsantander['purchase_price'] > 1000]
motosparsantander = motosparsantander.loc[motosparsantander['kilometers'] > 20]
motosparsantander['CODIGO DEALER'] = 'B67377580'
motosparsantander['NOMBRE DEALER'] = 'AJ MOTOR EUROPA, S.L.'
motosparsantander['PRODUCTO'] = 'O'
motosparsantander['NUM. OPERACION'] = 'ESET20225000402'
motosparsantander['BASTIDOR'] = motosparsantander['frame_number']
motosparsantander['IMPORTE'] = motosparsantander['purchase_price']
motosparsantander['MONEDA'] = 'EUR'
motosparsantander['MARCA'] = motosparsantander['brand']
motosparsantander['MODELO'] = motosparsantander['model']
motosparsantander['VERSION'] = motosparsantander.apply(lambda row: f"{row['MARCA']} {row['MODELO']}",axis=1)
motosparsantander['MATRICULA'] = motosparsantander['license_plate']
motosparsantander['FECHA MATRICULA'] = motosparsantander['registration_date']
motosparsantander['FACTURA'] = motosparsantander['purchase_id']
motosparsantander['FECHA FACTURA'] = motosparsantander['purchase_date']
motosparsantander = motosparsantander.sort_values(by='FECHA FACTURA',ascending=False)

#10 años
def filtrar_antiguedadsan(df, columna_fecha, anos_antiguedad):
    # Obtener la fecha actual
    fecha_actual = datetime.now().date()
    
    # Calcular la fecha límite
    fecha_limite = fecha_actual - timedelta(days=anos_antiguedad * 365)
    
    # Filtrar los datos donde la fecha de registro sea superior a la fecha límite
    datos_filtrados = df[pd.to_datetime(df[columna_fecha]).dt.date > fecha_limite]
    
    return datos_filtrados

motosparsantander = filtrar_antiguedadsan(motosparsantander, 'registration_date', 10)
columnas_doc_santander =[ 'CODIGO DEALER', 'NOMBRE DEALER', 'PRODUCTO', 'NUM. OPERACION', 'BASTIDOR', 'IMPORTE', 'MONEDA', 'MARCA', 'MODELO', 'VERSION', 'MATRICULA', 'FECHA MATRICULA', 'FACTURA', 'FECHA FACTURA']
motosparsantander = motosparsantander[columnas_doc_santander]
motosparsantander['stockapp'] = motosparsantander.apply(lambda row: f"{row['MATRICULA']}#santanderSales",axis=1)
motosparsantander = motosparsantander[~motosparsantander['MATRICULA'].isin(['7347MMT', '4624MNV', 'MAPK110', '2205LYR'])]
maxsantander = motosparsantander['IMPORTE'].sum()

#motoslibres para sabadell
motosparsabadell = motosparadoc.loc[motosparadoc['purchase_price'] > 2400]
motosparsabadell = motosparsabadell.loc[motosparsabadell['kilometers'] > 50]
motosparsabadell = motosparsabadell.sort_values(by='purchase_date',ascending=False)

#sabadell años
def filtrar_antiguedadsab(df, columna_fecha, años_min, años_max):
    # Obtener la fecha actual
    fecha_actual = datetime.now().date()
    
    # Calcular las fechas límite
    fecha_limite_min = fecha_actual - timedelta(days=años_max * 365)
    fecha_limite_max = fecha_actual - timedelta(days=años_min * 365)
    
    # Filtrar los datos donde la fecha de registro esté entre las dos fechas límite
    datos_filtrados = df[(pd.to_datetime(df[columna_fecha]).dt.date > fecha_limite_min) & 
                         (pd.to_datetime(df[columna_fecha]).dt.date <= fecha_limite_max)]

    return datos_filtrados

#motos para sabadell    
motosparsabadell = filtrar_antiguedadsab(motosparsabadell, 'registration_date', 10, 17)
motosparsabadell = pd.concat([motosparsabadell, extrasdesab], axis=0)
motosparsabadell = motosparsabadell.sort_values(by='purchase_date',ascending=False)
motosparsabadell = motosparsabadell.loc[motosparsabadell['sabadellsales'].isnull()]
motosparsabadell = motosparsabadell.loc[motosparsabadell['sofincosales'].isnull()]
motosparsabadell = motosparsabadell.drop_duplicates(subset=['license_plate'], keep='first')
motosparsabadell = motosparsabadell.loc[motosparsabadell['purchase_price'] > 2490]
motosparsabadell = motosparsabadell.loc[motosparsabadell['kilometers'] > 20]
motosparsabadell['Marca'] = motosparsabadell['brand']
motosparsabadell['Modelo'] = motosparsabadell['model']
motosparsabadell['Name'] = motosparsabadell.apply(lambda row: f"{row['Marca']} {row['Modelo']}",axis=1)
motosparsabadell['Matrícula'] = motosparsabadell['license_plate']
motosparsabadell['Nº Bastidor'] = motosparsabadell['frame_number']
motosparsabadell['kilometros'] = motosparsabadell['kilometers']
motosparsabadell['Año'] = motosparsabadell['registration_date'].dt.year
motosparsabadell['Precio compra'] = motosparsabadell['purchase_price']
motosparsabadell = motosparsabadell.sort_values(by='Precio compra',ascending=False)
columnas_doc_sabadeel = [ 'Name','Marca', 'Modelo', 'Matrícula', 'Nº Bastidor', 'kilometros', 'Año', 'Precio compra']
motosparsabadell = motosparsabadell[columnas_doc_sabadeel]
motosparsabadell['stockapp'] = motosparsabadell.apply(lambda row: f"{row['Matrícula']}#sabadellSales",axis=1)
maxsabadell = motosparsabadell['Precio compra'].sum()


#motos para wabi
motosparadocwabi = metabase.loc[metabase['stock_status'].isin(['rented'])]
motosparadocwabi = motosparadocwabi.loc[motosparadocwabi['actual_credit_policy'].isnull()]
motosparwabi = motosparadocwabi.loc[motosparadocwabi['santandersales'].isnull()]
motosparwabi = motosparwabi.loc[motosparwabi['wavi'].isnull()]
motosparwabi = motosparwabi.loc[motosparwabi['sabadellsales'].isnull()]
motosparwabi = motosparwabi.loc[motosparwabi['sofincosales'].isnull()]
motosparwabi = motosparwabi.loc[motosparwabi['purchase_price'] > 1000]
motosparwabi = motosparwabi.loc[motosparwabi['kilometers'] > 20]
motosparwabi['CODIGO DEALER'] = 'B67377580'
motosparwabi['NOMBRE DEALER'] = 'AJ MOTOR EUROPA, S.L.'
motosparwabi['PRODUCTO'] = 'O'
motosparwabi['NUM. OPERACION'] = 'ESET20235001800'
motosparwabi['BASTIDOR'] = motosparwabi['frame_number']
motosparwabi['IMPORTE'] = motosparwabi['purchase_price']
motosparwabi['MONEDA'] = 'EUR'
motosparwabi['MARCA'] = motosparwabi['brand']
motosparwabi['MODELO'] = motosparwabi['model']
motosparwabi['VERSION'] = motosparwabi.apply(lambda row: f"{row['MARCA']} {row['MODELO']}",axis=1)
motosparwabi['MATRICULA'] = motosparwabi['license_plate']
motosparwabi['FECHA MATRICULA'] = motosparwabi['registration_date']
motosparwabi['FACTURA'] = motosparwabi['purchase_id']
motosparwabi['FECHA FACTURA'] = motosparwabi['purchase_date']
motosparwabi = motosparwabi.sort_values(by='FECHA FACTURA',ascending=False)
motosparwabi = filtrar_antiguedadsan(motosparwabi, 'registration_date', 10)
motosparwabi = motosparwabi[columnas_doc_santander]
motosparwabi['stockapp'] = motosparwabi.apply(lambda row: f"{row['MATRICULA']}#wabi",axis=1)

def filtrar_facturacion(df, columna_fecha, mes_antiguedad):
    # Obtener la fecha actual
    fecha_actual = datetime.now().date()
    
    # Calcular la fecha límite
    fecha_limite = fecha_actual - timedelta(days=mes_antiguedad * 30)
    
    # Filtrar los datos donde la fecha de registro sea superior a la fecha límite
    datos_filtrados = df[pd.to_datetime(df[columna_fecha]).dt.date > fecha_limite]
    
    return datos_filtrados

#motos para sofinco
motosparsofinco = filtrar_facturacion(motosparsantander, 'FECHA FACTURA', 2)
motosparsofinco = motosparsofinco.merge(metabase[['frame_number', 'kilometers']], left_on='BASTIDOR', right_on='frame_number', how='left')
columnas_doc_sofinco = ['MARCA', 'MODELO', 'BASTIDOR', 'kilometers', 'MATRICULA', 'FECHA MATRICULA', 'FACTURA', 'IMPORTE']
motosparsofinco = motosparsofinco[columnas_doc_sofinco]
motosparsofinco['stockapp'] = motosparsofinco.apply(lambda row: f"{row['MATRICULA']}#sofincoSales",axis=1)
maxsofinco = motosparsofinco['IMPORTE'].sum()

#periodos
periodos = [0, 30, 60, 90, 120, 150, 180]

#Santander
Santandertot = Santander[Santander['Estado de la Ficha Téc.'].isin(['Recibida fotocopia' , 'Recibida' , 'Solicitada'])]
Santander = Santandertot.loc[Santandertot['Nº póliza'].isin([1019])]

totalporperiodosan = {}

for i in range(len(periodos) - 1):
    inicio_periodo = fecha_actual + timedelta(days=periodos[i])
    fin_periodo = fecha_actual + timedelta(days=periodos[i + 1])
    
    motos_en_periodo = Santander[
        (Santander['Fecha Vencimiento'] >= inicio_periodo) &
        (Santander['Fecha Vencimiento'] <= fin_periodo)
    ]
    
    total = motos_en_periodo['Importe Documentación'].sum()
    
    totalporperiodosan[f'Importe Docu ({periodos[i]}-{periodos[i+1]})'] = total

Santander =  pd.DataFrame.from_dict(totalporperiodosan, orient='index', columns=['Total'])

Santander = Santander.transpose()
Santander['Disponible'] = maxsantander
Santander['Póliza'] = 'Santander'

#Wabi
Wabi = Santandertot.loc[Santandertot['Nº póliza'].isin([1436])]

totalporperiodowab = {}

for i in range(len(periodos) - 1):
    inicio_periodo = fecha_actual + timedelta(days=periodos[i])
    fin_periodo = fecha_actual + timedelta(days=periodos[i + 1])
    
    motos_en_periodo = Wabi[
        (Wabi['Fecha Vencimiento'] >= inicio_periodo) &
        (Wabi['Fecha Vencimiento'] <= fin_periodo)
    ]
    
    total = motos_en_periodo['Importe Documentación'].sum()
    
    totalporperiodosan[f'Importe Docu ({periodos[i]}-{periodos[i+1]})'] = total

Wabi =  pd.DataFrame.from_dict(totalporperiodosan, orient='index', columns=['Total'])

Wabi = Wabi.transpose()
Wabi['Póliza'] = 'Wabi'

#Sabadell
totalporperiodosab = {}

for i in range(len(periodos) - 1):
    inicio_periodo = fecha_actual + timedelta(days=periodos[i])
    fin_periodo = fecha_actual + timedelta(days=periodos[i + 1])
    
    motos_en_periodo = Sabadell[
        (Sabadell['Fecha Vencimiento'] >= inicio_periodo) &
        (Sabadell['Fecha Vencimiento'] <= fin_periodo)
    ]
    
    total = motos_en_periodo['Importe Financiado'].sum()
    
    totalporperiodosab[f'Importe Docu ({periodos[i]}-{periodos[i+1]})'] = total

Sabadell =  pd.DataFrame.from_dict(totalporperiodosab, orient='index', columns=['Total'])

Sabadell = Sabadell.transpose()
Sabadell['Disponible'] = maxsabadell
Sabadell['Póliza'] = 'Sabadell'

#Sofinco
Sofinco = Sofinco[(Sofinco['Phase'] == 'Activo')]
totalporperiodosof = {}

for i in range(len(periodos) - 1):
    inicio_periodo = fecha_actual + timedelta(days=periodos[i])
    fin_periodo = fecha_actual + timedelta(days=periodos[i + 1])
    
    motos_en_periodo = Sofinco[
        (Sofinco['End date'] >= inicio_periodo) &
        (Sofinco['End date'] <= fin_periodo)
    ]
    
    total = motos_en_periodo['Amount'].sum()
    
    totalporperiodosof[f'Importe Docu ({periodos[i]}-{periodos[i+1]})'] = total

Sofinco =  pd.DataFrame.from_dict(totalporperiodosof, orient='index', columns=['Total'])

Sofinco = Sofinco.transpose()
Sofinco['Disponible'] = maxsofinco
Sofinco['Póliza'] = 'Sofinco'

#control
CreditStock = pd.concat([ Santander , Wabi, Sabadell , Sofinco], axis=0 , ignore_index=True)
CreditStockcolumnas = ['Póliza', 'Importe Docu (0-30)', 'Importe Docu (30-60)', 'Importe Docu (60-90)', 'Importe Docu (90-120)', 'Importe Docu (120-150)', 'Importe Docu (150-180)', 'Disponible']
CreditStock = CreditStock[CreditStockcolumnas]

#cambio formato final de fecha
metabase['purchase_date'] = pd.to_datetime(metabase['purchase_date']).dt.strftime('%d/%m/%Y')
metabase['registration_date'] = pd.to_datetime(metabase['registration_date']).dt.strftime('%d/%m/%Y')
Santanderp['Fecha Vencimiento'] = pd.to_datetime(Santanderp['Fecha Vencimiento']).dt.strftime('%d/%m/%Y')
Sabadellp['Fecha Vencimiento'] = pd.to_datetime(Sabadellp['Fecha Vencimiento']).dt.strftime('%d/%m/%Y')
Sofincop['End date'] = pd.to_datetime(Sofincop['End date']).dt.strftime('%d/%m/%Y')
Sofincop['Start date'] = pd.to_datetime(Sofincop['Start date']).dt.strftime('%d/%m/%Y')
rescate_santander['Fecha Vencimiento'] = pd.to_datetime(rescate_santander['Fecha Vencimiento']).dt.strftime('%d/%m/%Y')
motosparsantander['FECHA MATRICULA'] = pd.to_datetime(motosparsantander['FECHA MATRICULA']).dt.strftime('%d/%m/%Y')
motosparsantander['FECHA FACTURA'] = pd.to_datetime(motosparsantander['FECHA FACTURA']).dt.strftime('%d/%m/%Y')
motosparwabi['FECHA MATRICULA'] = pd.to_datetime(motosparwabi['FECHA MATRICULA']).dt.strftime('%d/%m/%Y')
motosparwabi['FECHA FACTURA'] = pd.to_datetime(motosparwabi['FECHA FACTURA']).dt.strftime('%d/%m/%Y')
motosparsofinco['FECHA MATRICULA'] = pd.to_datetime(motosparsofinco['FECHA MATRICULA']).dt.strftime('%d/%m/%Y')

#distribuir en hojas
with pd.ExcelWriter(archivo_final_excel, engine='xlsxwriter') as writer:
    # Escribir cada DataFrame en una hoja diferente
    metabase.to_excel(writer, sheet_name='Metabase', index=False)
    Santanderp.to_excel(writer, sheet_name='Santander', index=False)
    rescate_santander.to_excel(writer, sheet_name='R.Santander', index=False)
    motosparsantander.to_excel(writer, sheet_name='Motos Santander', index=False)
    motosparwabi.to_excel(writer, sheet_name='Motos Wabi', index=False)
    Sabadellp.to_excel(writer, sheet_name='Sabadell', index=False)
    rescate_sabadell.to_excel(writer, sheet_name='R.Sabadell', index=False)
    motosparsabadell.to_excel(writer, sheet_name='Motos Sabadell', index=False)
    Sofincop.to_excel(writer, sheet_name='Sofinco', index=False)
    rescate_sofinco.to_excel(writer, sheet_name='R.Sofinco', index=False)
    motosparsofinco.to_excel(writer, sheet_name='Motos Sofinco', index=False)
    CreditStock.to_excel(writer, sheet_name='Control', index=False)

# Abrir el archivo de Excel con openpyxl
wb = openpyxl.load_workbook(archivo_final_excel)

# Formatos y colores para cada hoja
sheet_formats = {
    'Metabase': {'color': 'FEFEFE', 'font_color': '509EE3', 'font_bold': True, 'font_name': 'Joanna Sans Nova Medium','cell_font_name': 'AwanZaman Medium','cell_font_size': 9}, #done
    'Santander': {'color': 'FFFFFF', 'font_color': 'EC0000', 'font_bold': True, 'font_name': 'Andante Display Demi Bold','cell_font_name': 'AwanZaman Medium','cell_font_size': 9}, #done
    'R.Santander': {'color': 'FFFFFF', 'font_color': 'EC0000', 'font_bold': True, 'font_name': 'Andante Display Demi Bold','cell_font_name': 'AwanZaman Medium','cell_font_size': 9}, #done
    'Sabadell': {'color': '028ACD', 'font_color': 'FFFFFF', 'font_bold': True, 'font_name': 'Sequel Sans Display Heavy','cell_font_name': 'AwanZaman Medium','cell_font_size': 9}, #done
    'R.Sabadell': {'color': '028ACD', 'font_color': 'FFFFFF', 'font_bold': True, 'font_name': 'Sequel Sans Display Heavy','cell_font_name': 'AwanZaman Medium','cell_font_size': 9}, #done
    'Sofinco': {'color': 'FFFFFF', 'font_color': '00A8B2', 'font_bold': True, 'font_name': 'Brahma Bold','cell_font_name': 'AwanZaman Medium','cell_font_size': 9}, #done
    'R.Sofinco': {'color': 'FFFFFF', 'font_color': '00A8B2', 'font_bold': True, 'font_name': 'Brahma Bold','cell_font_name': 'AwanZaman Medium','cell_font_size': 9}, #done
    'Motos Santander': {'color': 'FFFFFF', 'font_color': 'EC0000', 'font_bold': True, 'font_name': 'Andante Display Demi Bold','cell_font_name': 'AwanZaman Medium','cell_font_size': 9}, #done
    'Motos Sabadell': {'color': '028ACD', 'font_color': 'FFFFFF', 'font_bold': True, 'font_name': 'Sequel Sans Display Heavy','cell_font_name': 'AwanZaman Medium','cell_font_size': 9}, #done
    'Motos Wabi': {'color': 'FFFFFF', 'font_color': 'EC0000', 'font_bold': True, 'font_name': 'Andante Display Demi Bold','cell_font_name': 'AwanZaman Medium','cell_font_size': 9}, #done
    'Motos Sofinco': {'color': 'FFFFFF', 'font_color': '00A8B2', 'font_bold': True, 'font_name': 'Brahma Bold','cell_font_name': 'AwanZaman Medium','cell_font_size': 9}, #done
    'Control': {'color': 'EC0021', 'font_color': 'FFFFFF', 'font_bold': True,  'font_name': 'Artegra Sans Extended', 'italic': True,'cell_font_name': 'AwanZaman Medium','cell_font_size': 9} #done
}

#funcion para el fomrato
def format_sheet(sheet, fmt):
    # Eliminar las líneas de cuadrícula
    sheet.sheet_view.showGridLines = False
    
    # Crear formato de celda para títulos
    font_title = Font(bold=fmt['font_bold'], color=fmt['font_color'], name=fmt['font_name'], italic=fmt.get('italic', False))
    alignment = Alignment(horizontal="center", vertical="center")
    fill = PatternFill(start_color=fmt['color'], end_color=fmt['color'], fill_type="solid")

    # Crear formato de celda para celdas no titulares
    font_cell = Font(name=fmt['cell_font_name'], size=fmt['cell_font_size'])

    # Aplicar formato a los títulos de las columnas
    for col in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=1, column=col)
        cell.font = font_title
        cell.alignment = alignment
        cell.fill = fill

    # Ajustar el ancho de las columnas y aplicar formato a las celdas no titulares
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter  # Obtener la letra de la columna
        for cell in col:
            if cell.row != 1:
                cell.font = font_cell
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = max_length + 2
        sheet.column_dimensions[column].width = adjusted_width


# Aplicar formato a todas las hojas
for sheet_name, fmt in sheet_formats.items():
    sheet = wb[sheet_name]
    format_sheet(sheet, fmt)

# Guardar el archivo de Excel con los cambios
wb.save(archivo_final_excel)