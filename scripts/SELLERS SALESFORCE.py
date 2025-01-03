import pandas as pd
import openpyxl

MES= 11
AÑO = 2024

#Ruta archivo final excel
ruta_archivo_final_excel= "C:/Users/Ricardo Sarda/Desktop/MM/SELLERS/SELLERS.xlsx"

#Archivos
inf_usu_FC= "C:/Users/Ricardo Sarda/Downloads/ResumenCliente_20241112_210316.xlsx" #FC
inf_usu_AB= "C:/Users/Ricardo Sarda/Downloads/ResumenCliente_20241112_210704.xlsx" #AB
inf_usu_FT= "C:/Users/Ricardo Sarda/Downloads/ResumenCliente_20241112_210834.xlsx" #FT
archivo_ventas = "C:/Users/Ricardo Sarda/Downloads/FACTURADAS POR COMERCIAL-2024-11-13-12-14-28.xlsx" # Ruta del archivo del report de comerciales Solo detalles
archivo_leads= "C:/Users/Ricardo Sarda/Downloads/ES Sales - Leads por comercial meses-2024-11-13-12-15-05.xlsx" # Ruta del archivo leads Solo detalles
sellers_anterior = "C:/Users/Ricardo Sarda/Desktop/MM/SELLERS/SELLERS con mail.xlsx"
archivo_financiacion = "C:/Users/Ricardo Sarda/Downloads/FINANCIACIONES GENERAL 2024.xlsx" # Ruta del archivo de financiaciones

#Informes de usuario
#PBC
Limpiar_FC = openpyxl.load_workbook(inf_usu_FC)
hoja = Limpiar_FC.active
Limpiar_FC = pd.read_excel(inf_usu_FC)
Limpiar_FC = Limpiar_FC.loc[Limpiar_FC['SerieFactura'].isin(['FC','FP','FI','FL','AC'])]

Limpiar_AB = openpyxl.load_workbook(inf_usu_AB)
hoja = Limpiar_AB.active
Limpiar_AB = pd.read_excel(inf_usu_AB)

Limpiar_FT = openpyxl.load_workbook(inf_usu_FT)
hoja = Limpiar_FT.active
Limpiar_FT = pd.read_excel(inf_usu_FT)

SELLERS = pd.concat([Limpiar_FC, Limpiar_AB, Limpiar_FT])

SELLERSPBCpre = SELLERS[(SELLERS['FechaFactura'].dt.month == MES) & (SELLERS['FechaFactura'].dt.year == AÑO)]
SELLERSPBCpre = SELLERSPBCpre[SELLERSPBCpre['CodigoArticulo'].isin(['PACK COMPLETO' , 'PACK' , 'PACK PREMIUM' , 'STREET PLUS' , 'STREET 125' , 'STREET 300' , 'STREET 500' , 'SPORT PLUS' , 'SPORT 500' , 'SPORT 300'])]
columnas_seleccionadasPBC = ['SerieFactura' , 'NumeroFactura' , 'RazonSocial' , 'CodigoArticulo' , 'DescripcionArticulo', 'CodigoFamilia' , 'Unidades' , 'ImporteCoste', 'BaseImponible1', 'MargenBeneficio', 'PrecioCompra']
SELLERSPBCpre = SELLERSPBCpre[columnas_seleccionadasPBC]

SELLERSPBCpre = SELLERSPBCpre[(SELLERSPBCpre['SerieFactura'] == 'FT')]
SELLERSPBCpre = SELLERSPBCpre.loc[SELLERSPBCpre['CodigoFamilia'].isnull()]
SELLERSSPORT = SELLERSPBCpre[SELLERSPBCpre['CodigoArticulo'].isin(['SPORT PLUS' , 'SPORT 500' , 'SPORT 300'])]
SELLERSCOM = SELLERSPBCpre[SELLERSPBCpre['CodigoArticulo'].isin(['PACK COMPLETO' , 'PACK'])]
SELLERSPREM = SELLERSPBCpre[SELLERSPBCpre['CodigoArticulo'].isin(['PACK PREMIUM'])]
SELLERSSTREET = SELLERSPBCpre[SELLERSPBCpre['CodigoArticulo'].isin(['STREET PLUS' , 'STREET 125' , 'STREET 300' , 'STREET 500'])]

#LEADS
leads = pd.read_excel(archivo_leads)
leads['Fecha de creación'] = pd.to_datetime(leads['Fecha de creación'], format='%d/%m/%Y', dayfirst=True)
leads = leads[(leads['Fecha de creación'].dt.month == MES) & (leads['Fecha de creación'].dt.year == AÑO)]
leads = leads.groupby('Propietario del candidato').agg(
    leads=('Apellidos', 'count'),

).reset_index()

sellers_anterior = pd.read_excel(sellers_anterior)

leads = leads.merge(sellers_anterior[['Nombre', 'Vendedor']], left_on='Propietario del candidato', right_on='Nombre', how='left')

#B2C
SELLERSB2Cpre = SELLERS[(SELLERS['FechaFactura'].dt.month == MES) & (SELLERS['FechaFactura'].dt.year == AÑO)]
SELLERSB2Cpre = SELLERSB2Cpre[(SELLERSB2Cpre['SerieFactura'] == 'FC')]
SELLERSB2Cpre = SELLERSB2Cpre.loc[SELLERSB2Cpre['CodigoFamilia'].isnull()]
columnas_seleccionadasB2C = ['SerieFactura', 'NumeroFactura', 'RazonSocial', 'CodigoArticulo', 'DescripcionArticulo', 'BaseImponible1']
SELLERSB2Cpre = SELLERSB2Cpre[columnas_seleccionadasB2C]
SELLERSB2Cpre['CodigoArticulo'] = SELLERSB2Cpre['CodigoArticulo'].str.strip()
SELLERSB2Cpre['CodigoArticulo'] = SELLERSB2Cpre['CodigoArticulo'].astype(str)

#Report comerciales salesforce
ventas = pd.read_excel(archivo_ventas)
ventas['Matrícula'] = ventas['Nombre de la reserva'].str[-8:]
ventas['Matrícula'] = ventas['Matrícula'].str.replace(' ', '')

SELLERSB2C = SELLERSB2Cpre.merge(ventas[['Matrícula', 'Email Comercial']], left_on='CodigoArticulo', right_on='Matrícula', how='left')
SELLERSB2C = SELLERSB2C.merge(ventas[['Matrícula', 'Email Comercial']], left_on='CodigoArticulo', right_on='Matrícula', how='left')
SELLERSB2C = SELLERSB2C.merge(SELLERSCOM[['RazonSocial', 'Unidades']], left_on='RazonSocial', right_on='RazonSocial', how='left')
SELLERSB2C = SELLERSB2C.merge(SELLERSPREM[['RazonSocial', 'Unidades']], left_on='RazonSocial', right_on='RazonSocial', how='left')
SELLERSB2C = SELLERSB2C.merge(SELLERSSTREET[['RazonSocial', 'Unidades']], left_on='RazonSocial', right_on='RazonSocial', how='left')
SELLERSB2C = SELLERSB2C.merge(SELLERSSPORT[['RazonSocial', 'Unidades']], left_on='RazonSocial', right_on='RazonSocial', how='left', suffixes=('_SELLERSB2C', '_SELLERSSPORT'))

#Financiaciones

Financiacion = pd.read_excel(archivo_financiacion)

SELLERSB2C = SELLERSB2C.merge(Financiacion[['MATRÍCULA', 'IMPORTE FINANCIADO']], left_on='CodigoArticulo', right_on='MATRÍCULA', how='left')
SELLERSB2C = SELLERSB2C.merge(Financiacion[['MATRÍCULA', 'COMISIÓN']], left_on='CodigoArticulo', right_on='MATRÍCULA', how='left')

#Limpieza
SELLERSB2C.columns = SELLERSB2C.columns.str.replace('Unidades_x', 'Completo')
SELLERSB2C.columns = SELLERSB2C.columns.str.replace('Unidades_y', 'Premium')
SELLERSB2C.columns = SELLERSB2C.columns.str.replace('Unidades_SELLERSB2C', 'Street')
SELLERSB2C.columns = SELLERSB2C.columns.str.replace('Unidades_SELLERSSPORT', 'Sport')
SELLERSB2C.columns = SELLERSB2C.columns.str.replace('Email Comercial_x', 'Vendedor')
SELLERSB2C['MATRÍCULA_x'] = SELLERSB2C['MATRÍCULA_x'].apply(lambda x: 1 if pd.notna(x) and x != '' else x)
SELLERSB2C.columns = SELLERSB2C.columns.str.replace('MATRÍCULA_x', 'Finan')

columnas_seleccionadasB2Cfinales = ['SerieFactura', 'NumeroFactura', 'RazonSocial', 'CodigoArticulo', 'DescripcionArticulo', 'BaseImponible1' , 'Completo' , 'Premium' , 'Street' , 'Sport' , 'Finan' , 'Vendedor']
SELLERSB2C = SELLERSB2C[columnas_seleccionadasB2Cfinales]
SELLERSB2C = SELLERSB2C.drop_duplicates(subset='CodigoArticulo', keep='first')

SELLERSB2C['BaseImponible1'] = SELLERSB2C['BaseImponible1'].astype(float)
SELLERSB2C['Completo'] = SELLERSB2C['Completo'].astype(float)
SELLERSB2C['Premium'] = SELLERSB2C['Premium'].astype(float)
SELLERSB2C['Street'] = SELLERSB2C['Street'].astype(float)
SELLERSB2C['Sport'] = SELLERSB2C['Sport'].astype(float)
SELLERSB2C['Finan'] = SELLERSB2C['Finan'].astype(float)

#PIVOT
SUMMARYSELL = SELLERSB2C.groupby('Vendedor').agg(
    Unidades=('SerieFactura', 'count'),
    Facturación=('BaseImponible1', 'sum'),
    Financiación=('Finan', 'sum'),
    Completo = ('Completo', 'sum'),
    Premium = ('Premium', 'sum'),
    Street = ('Street', 'sum'),
    Sport = ('Sport', 'sum'),
).reset_index()


SUMMARY = SUMMARYSELL

SUMMARY['Unidades'] = SUMMARY['Unidades'].astype(float)
SUMMARY['Facturación'] = SUMMARY['Facturación'].astype(float)
SUMMARY['Financiación'] = SUMMARY['Financiación'].astype(float)
SUMMARY['Completo'] = SUMMARY['Completo'].astype(float)
SUMMARY['Premium'] = SUMMARY['Premium'].astype(float)
SUMMARY['Street'] = SUMMARY['Street'].astype(float)
SUMMARY['Sport'] = SUMMARY['Sport'].astype(float)

SUMMARY['Avg. Ticket'] = SUMMARY['Facturación'] / SUMMARY['Unidades']
SUMMARY['Finan %'] = (SUMMARY['Financiación'] / SUMMARY['Unidades'])
SUMMARY['PACKS %'] = (SUMMARY['Completo'] + SUMMARY['Premium']) / SUMMARY['Unidades']
SUMMARY['Kits %'] = (SUMMARY['Street'] + SUMMARY['Sport']) / SUMMARY['Unidades']
SUMMARY = SUMMARY.merge(leads[['Vendedor', 'leads']], left_on='Vendedor', right_on='Vendedor', how='left')

columnas_SUMMARY= ['Vendedor', 'Unidades', 'Facturación', 'Avg. Ticket' , 'Financiación' , 'Finan %' , 'Completo' , 'Premium' , 'PACKS %' , 'Street' , 'Sport' , 'Kits %' , 'leads']
SUMMARY = SUMMARY[columnas_SUMMARY]
SUMMARY.to_excel(ruta_archivo_final_excel , index = 'false')
