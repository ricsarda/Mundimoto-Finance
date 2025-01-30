import pdfplumber
import pandas as pd
import numpy as np
from datetime import datetime
from io import BytesIO
import re

def procesar_pdfs_en_memoria(pdfs_dict):
    """
    Primera parte:
    Procesa un diccionario {nombre_pdf: BytesIO} y convierte la información en un DataFrame
    con columnas: 'Fecha', 'Importe', 'Intereses', 'Total para Aplicar', 'Nuevo Capital Pendiente',
    'Codigo_Clave' y 'Nombre_Archivo'.
    """

    dataframes = []

    # Patrones de interés y su mapeo
    patrones_interes = ["Tfno", "EntregaImporte", "InteresesDevengados", "Totalpara", "NuevoCapital"]
    mapeo_nombres = {
        "Tfno": "Fecha",
        "EntregaImporte": "Importe",
        "InteresesDevengados": "Intereses",
        "Totalpara": "Total para Aplicar",
        "NuevoCapital": "Nuevo Capital Pendiente"
    }
    orden_nombres_final = list(mapeo_nombres.values())

    for pdf_name, pdf_buffer in pdfs_dict.items():
        try:
            with pdfplumber.open(pdf_buffer) as pdf:
                texto = "\n".join(
                    page.extract_text()
                    for page in pdf.pages
                    if page.extract_text()
                )

            # Extraer el código clave (p. ej. E31F...)
            codigo_clave = re.search(r'\bE\d{2}[A-Z]{1}\d{8,}\b', texto)
            codigo_clave_extraido = codigo_clave.group(0) if codigo_clave else None

            lineas = texto.split("\n")
            df_lineas = pd.DataFrame(lineas, columns=["Contenido"])

            # Filtrar las filas que empiezan con los patrones
            df_filtrado = df_lineas[df_lineas["Contenido"].str.startswith(tuple(patrones_interes), na=False)].copy()

            # Extraer la parte numérica al final de la línea
            df_filtrado["Datos"] = df_filtrado["Contenido"].str.extract(r'([\d.,/]+)$')
            df_filtrado["Contenido"] = df_filtrado["Contenido"].str.replace(r'([\d.,/]+)$', '', regex=True).str.strip()

            def cambiarnombres(texto_patron):
                for patron, nuevo_nombre in mapeo_nombres.items():
                    if texto_patron.startswith(patron):
                        return nuevo_nombre
                return texto_patron

            df_filtrado["Contenido"] = df_filtrado["Contenido"].apply(cambiarnombres)

            # Pivotar/transponer
            if df_filtrado.empty:
                # DataFrame vacío: creamos una fila vacía
                df_transpuesto = pd.DataFrame(columns=orden_nombres_final)
                df_transpuesto.loc[0] = [""] * len(orden_nombres_final)
            else:
                df_pivote = df_filtrado.set_index("Contenido")["Datos"]
                df_pivote = df_pivote.reindex(orden_nombres_final, fill_value="")
                df_transpuesto = df_pivote.to_frame().T.reset_index(drop=True)

            df_transpuesto["Operación"] = codigo_clave_extraido
            df_transpuesto["Archivo"] = pdf_name

            dataframes.append(df_transpuesto)

        except Exception as e:
            print(f"Error procesando {pdf_name}: {e}")

    if dataframes:
        df_final = pd.concat(dataframes, ignore_index=True)
    else:
        df_final = pd.DataFrame()

    return df_final

def main(files, pdfs, new_excel, month=None, year=None):
    """
    Función principal para “Financiaciones Renting”:
    1) Procesa PDFs (dict pdfs) y genera un DataFrame “Resumen”.
    2) Para cada PDF, extrae líneas [6:36], las traspone y las escribe en su propia hoja.
    3) Incrusta el código E... en A1 y la fecha recálculo en A2 de cada hoja.

    Retorna un BytesIO (new_excel) con el libro final.
    """
    try:
        # ------------------------------
        #     PRIMERA PARTE: RESUMEN
        # ------------------------------
        df_resultado = procesar_pdfs_en_memoria(pdfs)

        # Filtro de filas vacías en 'Fecha'
        if not df_resultado.empty and 'Fecha' in df_resultado.columns:
            df_resultado.dropna(subset=['Fecha'], inplace=True)
            df_resultado = df_resultado[df_resultado['Fecha'].str.strip() != '']

        # ------------------------------
        #     SEGUNDA PARTE: POR PDF
        # ------------------------------
        import openpyxl
        from openpyxl import Workbook

        # Creamos el libro final
        with pd.ExcelWriter(new_excel, engine='openpyxl') as writer:
            # Hoja “Resumen”
            if not df_resultado.empty:
                df_resultado.to_excel(writer, sheet_name="Resumen", index=False)
            else:
                # Si no hay nada, creamos una hoja vacía
                pd.DataFrame({"Mensaje":["No se encontraron datos en los PDFs"]}).to_excel(
                    writer, sheet_name='Resumen', index=False
                )

            # Recorremos cada PDF para extraer líneas [6:36]
            for pdf_name, pdf_buffer in pdfs.items():
                try:
                    with pdfplumber.open(pdf_buffer) as pdf_lectura:
                        texto = "\n".join(
                            page.extract_text()
                            for page in pdf_lectura.pages
                            if page.extract_text()
                        )

                    lineas = texto.split('\n')

                    # (1) Extraer código E... y fecha recálculo
                    match_codigo = re.search(r'\bE\d{2}[A-Z]\d{8,}\b', texto)
                    codigo_clave_extraido = match_codigo.group(0) if match_codigo else None

                    match_fecha = re.search(r"FECHARECALCULO\.\:\s+(\d{2}\/\d{2}\/\d{4})", texto)
                    fecha_recal = match_fecha.group(1) if match_fecha else None

                    # (2) Tomar líneas [6:36]
                    lineas_deseadas = lineas[6:36]
                    filas_spliteadas = []
                    for l in lineas_deseadas:
                        # Eliminar asterisco al inicio si lo hay
                        l = re.sub(r"^\*", "", l).strip()
                        columnas_splits = l.split()
                        filas_spliteadas.append(columnas_splits)

                    df_local = pd.DataFrame(filas_spliteadas)
                    if df_local.empty or df_local.shape[0] < 2:
                        # No podemos crear DF válido
                        continue

                    # (3) Asignar la 1ª fila como columnas
                    df_local.columns = df_local.iloc[0]
                    df_local = df_local.iloc[1:].reset_index(drop=True)

                    # Comprobamos si están las columnas “FECHA”, “CAPITAL”, “PENDIENTE”
                    columnas_necesarias = ['FECHA','CAPITAL','PENDIENTE']
                    if not set(columnas_necesarias).issubset(df_local.columns):
                        # Alguna columna no existe; skip
                        continue

                    # Filtramos solo esas 3
                    df_local = df_local[columnas_necesarias]

                    # Renombrar
                    df_local.rename(columns={
                        'FECHA': 'Fecha',
                        'CAPITAL': 'Amort anticipada',
                        'PENDIENTE': 'Fee'
                    }, inplace=True)

                    # Indicar que la columna 'Fecha' sea índice y transponer
                    df_local.set_index('Fecha', inplace=True)
                    df_local = df_local.T

                    df_local.dropna(axis=1, how='all', inplace=True)

                    # (4) Nombre de hoja (máx 31 chars)
                    nombre_hoja = pdf_name[:31]

                    # (5) Guardar en el Excel
                    df_local.to_excel(writer, sheet_name=nombre_hoja, startrow=3, startcol=1, index=False)

                    # (6) Insertar código y fecha en celdas (A1, A2)
                    ws = writer.book[nombre_hoja]
                    ws["A1"] = codigo_clave_extraido if codigo_clave_extraido else "COD NO ENCONTRADO"
                    ws["A2"] = fecha_recal if fecha_recal else "FECHA NO ENCONTRADA"

                    # Títulos "Amort anticipada" y "Fee" en A5/A6 (opcional)
                    ws["A5"] = "Amort anticipada"
                    ws["A6"] = "Fee"

                except Exception as e:
                    print(f"Error al procesar PDF {pdf_name} en segunda parte: {e}")

        # Retornar el archivo Excel en memoria
        new_excel.seek(0)
        return new_excel

    except Exception as e:
        raise RuntimeError(f"Error al procesar el script: {str(e)}")
